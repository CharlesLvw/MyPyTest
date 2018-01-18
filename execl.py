import xlrd
import openpyxl
import time
import threading




def Test():
    count = 1
    Count = 85
    path = r'C:\Users\17763\Desktop\test\test123.xls'
    openfile = xlrd.open_workbook(path)  # 打开
    sheets = openfile.sheet_by_name(u'Sheet1')  # 读取sheet
    hangshu = sheets.nrows
    # lieshu = sheets.ncols
    wb = openpyxl.Workbook()
    # print(wb.sheetnames)
    ws = wb.active
    # ws = wb.create_sheet("test1")
    ws.title = "test"
    try:
        for i in range(hangshu):
            rel = sheets.row_values(i)[5]
            # print(len(rel))
            print(count,sheets.row_values(rowx=count))

            if len(rel) != 18:
                count +=1
            else:
                trel = sheets.row_values(i)[5][8:10]
                print(trel)
                that = trel[0]
                if that == 0:
                    count +=1
                else:
                    if int(trel) >= Count:
                        # print(rel)
                        need = sheets.row_values(rowx=count) #获取当前行
                        # print(need)
                        ws.append(need)
                        wb.save(r'C:\Users\17763\Desktop\test\ttt.xlsx')
                        # print(need)
                        count +=1
                        time.sleep(1)
                    else:
                        count +=1
                        time.sleep(1)
    except IndexError:
        pass

Test()
